import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

class SaveToExcel:
    def save(self, array_in, filename, title_chart, row_range):
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active

        # Write data to the Excel sheet
        for row in array_in:
            sheet.append(row)

        # Create a LineChart object
        chart = LineChart()

        # Define the data range for the chart
        values = Reference(sheet, min_col=2, min_row=1, max_col=5, max_row=row_range)
        categories_date = Reference(sheet, min_col=1, min_row=2, max_row=row_range)

        # Add data to the chart
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories_date)

        # Set chart title and axis labels
        chart.title = title_chart
        chart.x_axis.title = "Date_Time"
        chart.y_axis.title = "Value"

        # Set the style of the chart
        chart.style = 12

        chart.x_axis.crosses = "min"
        chart.width = 15  # Width in character units
        chart.height = 10  # Height in character units

        # Add the chart to the worksheet at cell E3
        sheet.add_chart(chart, "G2")

        # Save the workbook
        workbook.save(filename)

        # ================== Convert str to float in excel =====================
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Iterate over the specified range (B2 to D6)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=row_range, min_col=2, max_col=5, values_only=True), start=2):
            for col_idx, cell in enumerate(row, start=2):
                # Check if the cell value is text
                if isinstance(cell, str):
                    # Attempt to convert the text to a decimal number
                    try:
                        cell_value = float(cell)
                        # Update the cell value with the converted decimal number
                        sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    except ValueError:
                        # Ignore if the conversion fails
                        pass

        # Save the modified Excel file
        workbook.save(filename)
